# -*- coding: utf-8 -*-
from openpyxl.workbook import Workbook as Workbook
from openpyxl.reader.excel import load_workbook as load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
from config_module import path_to_data_dir, path_to_data, path_to_final_exel_file


class ReportModule():


    def __init__(self) -> None:
        self._path_to_data_dir = path_to_data_dir
        self._path_to_data = path_to_data
        self._path_to_final_exel_file = path_to_final_exel_file
        return None
    

    def _make_book(self, path_to_data) -> None:
        '''
        Создаем xlsx файл для записи в него
        '''
        wb = Workbook()                         #Создает exel file
        wb.save(path_to_data)                   # сохнаняем с именем лежащей в переменной
        return None


    def _make_dirs(self, path_to_data_dir) -> None:
        '''
        Создаем папки для рабочих 'ДЛЯ УДОБСТВА' процессов
        '''
        os.makedirs(path_to_data_dir)
        return None
    

# region important
    #def add_style():
    #    '''
    #    Работает
    #    '''
    #    #from openpyxl import Workbook
    #    thin_border = Border(left=Side(style='thin'), 
    #                         right=Side(style='thin'), 
    #                         top=Side(style='thin'), 
    #                         bottom=Side(style='thick'))  #жирная полоса
    #    wb = Workbook()
    #    ws = wb.active
    #    ws.cell(row=3, column=2).border = thin_border
    #    wb.save('border_test.xlsx')
#endregion


    def _top_matrix_to_file(self, path_to_file) -> None:
        list = ['ИНН', 'Индекс формы', 'Наименование формы', 'Периодичность формы', 'Срок сдачи формы', 
                'Отчетный период','Комментарий', 'ОКУД', 'Дата актуализации перечня форм']  
        #wb = Workbook()
        wb = load_workbook(filename= path_to_file)
        ws = wb.active
        ws.append(list)
        wb.save(path_to_file)
        wb.close
        return None


    def read_exel_inn(self)->int:
        _list_inn = []
        book = load_workbook(self._path_to_data, read_only=True)
        sheet = book.active
        for row in sheet.iter_rows(min_col = 1, max_col=1, min_row=1, max_row = sheet.max_row):
            #nim_row минимальное количество строк 
            #max_col максимальное количество столбцов 
            #max_row максильманое колличество строк
            for data in row:
                inn = data.value
                if type(inn) is int and inn not in _list_inn:
                    _list_inn.append(inn)
                    yield inn


    @property
    def checking_existence_files(self) -> None:
        ''' 
        Данный блок проверяет создан ли файл базы данных
        если нет, то просто создает его
        таким образом мы пытаемся обойти ошибки со стороны 
        обработки
        '''

        '''
        Созадаем дерикторию
        '''
        if not os.path.exists(self._path_to_data_dir):               # Проверяем есть ли ДИРЕКТОРИЯ с таким именем, если нет, создает
            self._make_dirs(self._path_to_data_dir)

        '''
        Создаем базу данных в xlsx формате
        '''
        if not os.path.exists(self._path_to_data):                   # проверяет существование пути, если нет, вызываем функцию создания
            self._make_book(self._path_to_data)   

        '''Сюда нужно еще написать вызов функции которая вызывает окно пользователя и говорит ему
        что база пустая и необходимо ее пополнить INN для прохождения по ней
        '''
        if not os.path.exists(self._path_to_final_exel_file):
            self._make_book(self._path_to_final_exel_file)
            self._top_matrix_to_file(self._path_to_final_exel_file)
        print('OK')
        return None
    
    
    @property
    def _formater_to_exel(self) -> None:
        wb = load_workbook(self._path_to_final_exel_file)
        ws = wb.active
        #print(ws.max_row)
        '''
        Делаем фиксированный размер колонок по ширине
        '''
        ws.column_dimensions['A'].width = 12    
        ws.column_dimensions['B'].width = 30    
        ws.column_dimensions['C'].width = 40    
        ws.column_dimensions['D'].width = 15    
        ws.column_dimensions['E'].width = 40    
        ws.column_dimensions['F'].width = 15     
        ws.column_dimensions['G'].width = 65    
        ws.column_dimensions['H'].width = 15    
        ws.column_dimensions['I'].width = 15    

        ''' 
        первую строку пропускаем,
        узнаем максимальное колличество строк в документе
        '''
        i = 1
        while i <= ws.max_row:                                                                           #max_row - узнаем максимальное колличество строк
            ws[f'A{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    #Выравнивание текста по центру и перенос текста True
            ws[f'B{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'C{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'D{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'E{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'F{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'G{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'H{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)    
            ws[f'I{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            '''
            делаем выравнивание по высоте,
            т.к G ячейка самая большая, ориентируюсь на нее
            беру значение длинны списка и делю на /3
            '''
            u = len(ws[f'G{i}'].value)
            level = int(u/3)

            '''
            минимально допустимая высота ячейки
            (сделанна для 1о строчных клеток)
            '''
            if level < 60:          
                level = 60
            ws.row_dimensions[i].height = level                                                                 #размер строки    
            i+=1

        wb.save(self._path_to_final_exel_file)
        wb.close
        return None
    

    def writer_a_report_file(self, data:list) -> None:
        wb = load_workbook(self._path_to_final_exel_file)
        ws = wb.active
        ws.append(data)
        wb.save(self._path_to_final_exel_file)
        wb.close
        self._formater_to_exel
        return None
      