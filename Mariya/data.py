import os.path
from openpyxl.workbook import Workbook as Workbook
from openpyxl.reader.excel import load_workbook as load_workbook
from openpyxl.styles.borders import Border, Side
import os
import time
from driver import auto_page_pass
from config import(path_to_data_dir, path_to_data, final_exel_file, path_to_loads_dir)


def what_in_folder(path_to_request_dir):
    path_to_dir = os.listdir(path_to_request_dir)
    #print(f"список список файлов, из {path_to_request_dir} >> {path_to_dir}")
    return path_to_dir


def make_book(path_to_data):
    '''
    Создаем xlsx файл для записи в него
    '''
    wb = Workbook()
    wb.save(path_to_data)     # сохнаняем с именем лежащей в переменной


def make_dirs(path_to_data_dir):
    '''
    Создаем папки для рабочих 'ДЛЯ УДОБСТВА' процессов
    '''
    os.makedirs(path_to_data_dir)


def read_dataFile2():
    '''
    Построчное чтение xlsx
    '''
    book = load_workbook(path_to_data, read_only=True)
    sheet = book["Sheet"]
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=999):
        '''
        nim_row минимальное количество строк 
        max_col максимальное количество столбцов 
        max_row максильманое колличество строк
        '''
        for data_list in row:
            #print(data_list)
            data_list = data_list.value
            print(data_list)
            #auto_page_pass(data_list)
#read_dataFile2()

def read_dataFile22():
    '''
    Построчное чтение xlsx
    '''
    wb = Workbook()
    book = load_workbook(test_path, read_only=True)
    #sheet = book["Sheet1"]
    #sheet1 = book.active
    #print(sheet1)
    #print(book.get_sheet_names()) #проверка списка страниц
    i = 2
    sheet = book.active                                                                 #открываем активную страницу
    print(sheet.max_row)                                                                #по сути работает как len строк
    print(sheet.max_column)                                                             #по сути работает как len столбцов
    for row in sheet.iter_rows(min_col = 1,  min_row=i, max_col=9, max_row=i):
        '''
        nim_row минимальное количество строк 
        max_col максимальное количество столбцов 
        max_row максильманое колличество строк
        '''
        for data in row:
            #print(data_list)
            out_data = data.value
            print(out_data)
            #auto_page_pass(data_list)
#read_dataFile22()


def read_exel_inn(path, min_and_max_col):
    list = []
    i = min_and_max_col
    book = load_workbook(path, read_only=True)
    sheet = book.active
    for row in sheet.iter_rows(min_col = i, max_col=i, min_row=1, max_row = sheet.max_row):
        #nim_row минимальное количество строк 
        #max_col максимальное количество столбцов 
        #max_row максильманое колличество строк
        for data in row:
            out = data.value
            if type(out) is int and out not in list:
                list.append(out)
            else:
                break
    for u in list:
        auto_page_pass(u)
    #print(list)
    #auto_page_pass().browser.quit()
    return None

#read_exel_inn(path_to_home_dir + '\\data\\dataNEW.xlsx', 1)

def read_exel(path, min_and_max_col):
    list = []
    i = min_and_max_col
    book = load_workbook(path, read_only=True)
    sheet = book.active
    for row in sheet.iter_rows(min_col = 1, max_col=i, min_row=1, max_row = sheet.max_row):
        #nim_row минимальное количество строк 
        #max_col максимальное количество столбцов 
        #max_row максильманое колличество строк
        for data in row:
            out = data.value
            out1 = str(data)
            if out not in list:
                list.append(out)
            else:
                break
    print(list)
    return list

def append_in_data(list, path_to_file):
    wb = Workbook()
    ws = wb.active
    #print(value)
    for item in list:
        ws.append(item)
    wb.save(path_to_file)
    return(print('append in data >> ok'))

def list_in_list(list):
    '''
    ОБЕРТКА СПИСОК В СПИСОК list[list[]]
    '''
    total_list = []
    for i in list:
        u = [i]
        total_list.append(u)
    return total_list
 

def add_style():
    '''
    Работает
    '''
    #from openpyxl import Workbook
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thick'))  #жирная полоса
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=2).border = thin_border
    wb.save('border_test.xlsx')


#(read_exel_inn(path_to_data))
#append_in_data2(list_in_list(read_exel_inn(test_path)))
#append_in_data(list_in_list(read_exel_inn(path_to_home_dir + '\\data\\Список дочек для статистики.xlsx', 4)), path_to_home_dir + '\\data\\dataNEW.xlsx')
#read_exel_inn(path_to_home_dir + '\\data\\Список дочек для статистики.xlsx', 4)


def passage_in_folder_files(path_to_dir):
    folder = what_in_folder(path_to_dir)
    i = '\\'
    data = read_exel(path_to_dir + i + folder[0], 11)
    ##for item in folder:
    ##    i = '\\'
    ##    #print(item)
    ##    data = read_exel(path_to_dir + i + item, 11)
    ##    print(data)
    ###print(folder)
    ##print(path_to_dir + i + item)
    for item in data:
        print(data.index(item), item)

#passage_in_folder_files(loads_dir)



def exel_reader(path_to_file) -> list:
    book = load_workbook(path_to_file, read_only=True)
    sheet = book.active
    ws = sheet['A1' : 'K16']
    list = []
    for row in ws:
        for cell in row:
                out = cell.value
                if out is not None:
                    list.append(out)
                    #print(list.index(out), out)
                else:
                    break

    return(list)
 

def example(list):
    for item in list:
        index = list.index(item)
        print(f'{index}>> {item}')

def writer_a_report_file(list, path_to_file):
    #wb = Workbook()                                   #перезаписывает файл
    wb = load_workbook(path_to_file)                   #Добавляет в файл 
    ws = wb.active
    #print(value)
    for item in list:
        ws.append(item)
    #ws.append(list)
    wb.save(path_to_file)
    return(print(f'writer_a_report_file >> ok'))

def top_matrix_to_file(path_to_file) -> list:
    list = ['ИНН', 'Индекс формы', 'Наименование формы', 'Периодичность формы', 'Срок сдачи формы', 
             'Отчетный период','Комментарий', 'ОКУД', 'Дата актуализации перечня форм']
            
    #wb = Workbook()
    wb = load_workbook(path_to_file)
    ws = wb.active
    ws.append(list)
    wb.save(path_to_file)
    wb.close

#top_matrix_to_file(test)

def matrix_to_file(list) -> list:
    #for item in list:
    #    print(list.index(item) ,item)
    #принимает только exel_reader 
    
        #ЕСЛИ ДЛИННА СПИСКА ЮОЛЬШЕ ЧЕМ СТОЛЬКО ТО, ДЕЛАЕМ ВОТ ЭТО
 
        if len(list) > 100:
             list1 = [
                [list[17], list[32], list[33], list[34], list[35], list[36], list[37], list[38], list[16]],         #1
                [list[17], list[39], list[40], list[41], list[42], list[43], list[44], list[45], list[16]],         #2
                [list[17], list[46], list[47], list[48], list[49], list[50], list[51], list[52], list[16]],         #3
                [list[17], list[53], list[54], list[55], list[56], list[57], list[58], list[59], list[16]],         #4
                [list[17], list[60], list[61], list[62], list[63], list[64], list[65], list[66], list[16]],         #5
                [list[17], list[67], list[68], list[69], list[70], list[71], list[72], list[73], list[16]],         #6
                [list[17], list[74], list[75], list[76], list[77], list[78], list[79], list[80], list[16]],         #7
                [list[17], list[81], list[82], list[83], list[84], list[85], list[86], list[87], list[16]],         #8
                [list[17], list[88], list[89], list[90], list[91], list[92], list[93], list[94], list[16]],         #9
                [list[17], list[95], list[96], list[97], list[98], list[99], list[100], list[101], list[16]],       #10
                    ]
             return(list1)

        elif len(list) > 93:
            list2 = [
                [list[17], list[32], list[33], list[34], list[35], list[36], list[37], list[38], list[16]],         #1
                [list[17], list[39], list[40], list[41], list[42], list[43], list[44], list[45], list[16]],         #2
                [list[17], list[46], list[47], list[48], list[49], list[50], list[51], list[52], list[16]],         #3
                [list[17], list[53], list[54], list[55], list[56], list[57], list[58], list[59], list[16]],         #4
                [list[17], list[60], list[61], list[62], list[63], list[64], list[65], list[66], list[16]],         #5
                [list[17], list[67], list[68], list[69], list[70], list[71], list[72], list[73], list[16]],         #6
                [list[17], list[74], list[75], list[76], list[77], list[78], list[79], list[80], list[16]],         #7
                [list[17], list[81], list[82], list[83], list[84], list[85], list[86], list[87], list[16]],         #8
                [list[17], list[88], list[89], list[90], list[91], list[92], list[93], list[94], list[16]],         #9
                #[list[17], list[95], list[96], list[97], list[98], list[99], list[100], list[101], list[16]],      #10
                    ]
            return(list2)

        elif len(list) > 86:
            list3 = [
                [list[17], list[32], list[33], list[34], list[35], list[36], list[37], list[38], list[16]],         #1
                [list[17], list[39], list[40], list[41], list[42], list[43], list[44], list[45], list[16]],         #2
                [list[17], list[46], list[47], list[48], list[49], list[50], list[51], list[52], list[16]],         #3
                [list[17], list[53], list[54], list[55], list[56], list[57], list[58], list[59], list[16]],         #4
                [list[17], list[60], list[61], list[62], list[63], list[64], list[65], list[66], list[16]],         #5
                [list[17], list[67], list[68], list[69], list[70], list[71], list[72], list[73], list[16]],         #6
                [list[17], list[74], list[75], list[76], list[77], list[78], list[79], list[80], list[16]],         #7
                [list[17], list[81], list[82], list[83], list[84], list[85], list[86], list[87], list[16]],         #8
                #[list[17], list[88], list[89], list[90], list[91], list[92], list[93], list[94], list[16]],        #9
                #[list[17], list[95], list[96], list[97], list[98], list[99], list[100], list[101], list[16]],      #10
                    ]
            return(list3)

        elif len(list) > 79:
            list4 = [
                [list[17], list[32], list[33], list[34], list[35], list[36], list[37], list[38], list[16]],         #1
                [list[17], list[39], list[40], list[41], list[42], list[43], list[44], list[45], list[16]],         #2
                [list[17], list[46], list[47], list[48], list[49], list[50], list[51], list[52], list[16]],         #3
                [list[17], list[53], list[54], list[55], list[56], list[57], list[58], list[59], list[16]],         #4
                [list[17], list[60], list[61], list[62], list[63], list[64], list[65], list[66], list[16]],         #5
                [list[17], list[67], list[68], list[69], list[70], list[71], list[72], list[73], list[16]],         #6
                [list[17], list[74], list[75], list[76], list[77], list[78], list[79], list[80], list[16]],         #7
                #[list[17], list[81], list[82], list[83], list[84], list[85], list[86], list[87], list[16]],        #8
                #[list[17], list[88], list[89], list[90], list[91], list[92], list[93], list[94], list[16]],        #9
                #[list[17], list[95], list[96], list[97], list[98], list[99], list[100], list[101], list[16]],      #10
                ]
            return(list4)

        else:
            return(print('Выход за диапазон. Обратитесь к Разработчику.'))




#writer_a_report_file(matrix_to_file(exel_reader('C:\\Users\\NZXT\\source\\repos\\DELAGREEN\\Mariya\\data\\loads\\Сведения о кодах и формах (1).xlsx')), test) 
# Прикол какой, лист в лист нужно поменять и сделать такой как матрица которую я составил
#example(exel_reader('C:\\Users\\NZXT\\source\\repos\\DELAGREEN\\Mariya\\data\\loads\\Сведения о кодах и формах (1).xlsx'))


def final_define_write_a_report_file(path_to_dir, path_to_file):
    folder = what_in_folder(path_to_dir)
    i = '\\'
    for item in folder:
        print(item)
        print(path_to_dir + i + item)
        writer_a_report_file(matrix_to_file(exel_reader(path_to_dir + i + item)), path_to_file)
        #writer_a_report_file(exel_reader(path_to_dir + i + item), path_to_file)
    print('Complite')
    

#final_define_write_a_report_file(path_to_loads_dir, final_exel_file)


#example(exel_reader(test))

''' 
Данный блок проверяет создан ли файл базы данных
если нет, то просто создает его
таким образом мы пытаемся обойти ощибки со стороны 
обработки программой
'''
'''
Созадаем дерикторию
'''
if not os.path.exists(path_to_data_dir):             # Проверяем есть ли ДИРЕКТОРИЯ с таким именем, если нет, создает
    make_dirs(path_to_data_dir)
'''
Создаем базу данных в xlsx формате
'''
if not os.path.exists(path_to_data):            # проверяет существование пути, если нет, вызываем функцию создания
    make_book(path_to_data)                             
                                                # Сюда нужно еще написать вызов функции которая вызывает окно пользователя и говорит ему
                                                # что база пустая и необходимо ее пополнить INN для прохождения по ней
if not os.path.exists(final_exel_file):
    make_book(final_exel_file)
    top_matrix_to_file(final_exel_file)